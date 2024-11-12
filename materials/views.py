from rest_framework import viewsets, status
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework.views import APIView
from openpyxl import load_workbook
from materials.models import Material, Category
from materials.serializers import MaterialSerializer, CategorySerializer
from typing import Optional, Dict, Any, List


class MaterialViewSet(viewsets.ModelViewSet):
    """ ViewSet для модели Material, поддерживающий стандартные CRUD операции """

    queryset = Material.objects.all()
    serializer_class = MaterialSerializer


class CategoryViewSet(viewsets.ModelViewSet):
    """ ViewSet для модели Category, поддерживающий стандартные CRUD операции """

    queryset = Category.objects.all()
    serializer_class = CategorySerializer


class CategoryTreeAPIView(APIView):
    """
    APIView для получения иерархического дерева категорий и их материалов,
    с вычислением общей стоимости по каждой категории
    """

    def get(self, request) -> Response:
        def get_category_tree(category: Category) -> Dict[str, Any]:
            """ Рекурсивная функция для построения дерева категорий """

            children = [get_category_tree(child) for child in category.children.all()]
            materials = Material.objects.filter(category=category)
            total_price = sum(material.price for material in materials) + sum(child['total_price'] for child in children)

            return {
                'id': category.id,
                'name': category.name,
                'code': category.code,
                'materials': MaterialSerializer(materials, many=True).data,
                'children': children,
                'total_price': total_price
            }

        root_categories = Category.objects.filter(parent=None)
        category_tree = [get_category_tree(category) for category in root_categories]
        return Response(category_tree, status=status.HTTP_200_OK)


@api_view(['POST'])
def upload_materials(request) -> Response:
    """ Загрузка данных о категориях и материалах из Excel файла """

    file = request.FILES.get('file')
    if not file:
        return Response({'error': 'Файл не предоставлен'}, status=status.HTTP_400_BAD_REQUEST)
    if not file.name.endswith('.xlsx'):
        return Response({'error': 'Неверный формат файла. Требуется .xlsx'}, status=status.HTTP_400_BAD_REQUEST)

    wb = load_workbook(filename=file, data_only=True)

    # Обработка листа "Категории"
    if 'Категории' in wb.sheetnames:
        ws_categories = wb['Категории']
        for row in ws_categories.iter_rows(min_row=2, max_col=3, values_only=True):

            # Пропуск строки, если она полностью пустая
            if all(cell is None for cell in row):
                continue

            name, parent_name, code = row

            # Проверка наличия обязательных данных
            if not all([name, code]):
                print(f"Пропуск строки категории из-за пропущенных значений: {row}")
                continue
            parent_category = Category.objects.filter(name=parent_name).first() if parent_name else None

            # Проверка существования категории с таким именем
            try:
                category = Category.objects.get(name=name)

            # Если категория не существует, создаётся новая
            except Category.DoesNotExist:
                category = Category.objects.create(
                    name=name,
                    code=code,
                    parent=parent_category
                )

    # Обработка листа "Материалы"
    if 'Материалы' in wb.sheetnames:
        ws_materials = wb['Материалы']
        for row in ws_materials.iter_rows(min_row=2, max_col=4, values_only=True):
            name, category_name, code, price = row

            # Проверка наличия обязательных данных
            if not all([name, category_name, code, price]):
                print(f"Пропуск строки категории из-за пропущенных значений: {row}")
                continue

            category = Category.objects.filter(name=category_name).first()
            try:
                price = float(price)
            except (TypeError, ValueError):
                print(f"Пропуск строки категории из-за пропущенных значений: {row}")
                continue

            Material.objects.update_or_create(
                code=code,
                defaults={'name': name, 'category': category, 'price': price}
            )

    return Response({'status': 'Файл успешно загружен'}, status=status.HTTP_201_CREATED)




